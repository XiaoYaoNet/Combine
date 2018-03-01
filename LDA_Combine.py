import xlrd
import xlwt
import openpyxl
import string
import sys

Pos_Table1 = sys.argv[1]
Pos_Table2 = sys.argv[2]
Pos_Table_End=sys.argv[2]
Count1=int(sys.argv[3])-1
Count2=int(sys.argv[4])-1

def read07Excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name('Sheet1')

    sheet_value=[]
    for row in sheet.rows:
        tmp=[]
        for cell in range(len(row)):
            tmp.append(row[cell].value)
        sheet_value.append(tmp)
    return sheet_value

def write07Excel(path,value):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    wb.save(path)
    print("写入数据成功！")

Table1_txt=read07Excel(Pos_Table1)
Table2_txt=read07Excel(Pos_Table2)

total=[]
for i in range(len(Table1_txt)):
    tmp=[]
    for j in range(len(Table2_txt)):
    	if str(Table1_txt[i][Count1])==str(Table2_txt[j][Count2]):
            for x in range(len(Table1_txt[i])):
                if x==Count1:
                    continue
                Table2_txt[j].append(Table1_txt[i][x])
            tmp=Table2_txt[j]
            total.append(tmp)
            break
write07Excel(Pos_Table_End,total)